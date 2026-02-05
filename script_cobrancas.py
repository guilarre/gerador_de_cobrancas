# Script para gerar cobranças

from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

file_path = 'Controle pagamentos - aulas de inglês.xlsx'

# Dataframe geral com todos os worksheets
df = pd.read_excel(file_path, sheet_name=None, decimal=',')

# Lista com cada sheet como dataframe
lista_alunos = []

# Criando uma lista apenas das planilhas que estão visíveis
# (i.e. não ocultas no excel)
# Obs. tem que usar o load_workbook do openpyxl pra poder
# verificar visibilidade
wb = load_workbook(file_path)
visible_sheets = [sheet.title for sheet in wb.worksheets if sheet.sheet_state == 'visible']

# Criar lista de alunos como dicionários
sheet_count = 0
for sheet in df.keys():
    if sheet in visible_sheets:
        df_sheet = pd.read_excel(
            'Controle pagamentos - aulas de inglês.xlsx', sheet_name=sheet)

        aluno = {}

        nome = df_sheet['Aluno'][0]
        aluno['Nome'] = nome

        aluno['Datas'] = []
        for data in df_sheet['Data']:
            aluno['Datas'].append(data)

        aluno['Horas'] = []
        for horas in df_sheet['Horas de aula']:
            aluno['Horas'].append(horas)

        aluno['Preco'] = df_sheet['Hora/aula (R$)'][0]

        lista_alunos.append(aluno)
        sheet_count += 1


# Função para retornar mês passado
def month_from_number(month_number):
    if month_number == 1:
        return "Janeiro"
    elif month_number == 2:
        return "Fevereiro"
    elif month_number == 3:
        return "Março"
    elif month_number == 4:
        return "Abril"
    elif month_number == 5:
        return "Maio"
    elif month_number == 6:
        return "Junho"
    elif month_number == 7:
        return "Julho"
    elif month_number == 8:
        return "Agosto"
    elif month_number == 9:
        return "Setembro"
    elif month_number == 10:
        return "Outubro"
    elif month_number == 11:
        return "Novembro"
    elif month_number == 12:
        return "Dezembro"


last_month_number = (datetime.now() - timedelta(days=30)).month
last_month_name = month_from_number(last_month_number)


# Função pra arredondar número apenas se tiver .0 como decimal
def conditional_round(number):
    try:
        if number == int(number):
            return int(number)
        else:
            return number
    except ValueError:
        pass


# Função pra retornar cumprimento (bom dia, boa tarde ou boa noite)
def cumprimento():
    hora_atual = datetime.now().hour

    if 0 < hora_atual < 12:
        return "Bom dia"
    elif 12 <= hora_atual < 18:
        return "Boa tarde"
    elif 18 <= hora_atual < 23:
        return "Boa noite"


# Resetando cobrancas.txt:
p = Path(__file__).with_name('cobrancas.txt')
with p.open('w', encoding='utf-8') as file:
    pass


# Escrevendo cobrancas.txt
for aluno in lista_alunos:
    nome = aluno['Nome']
    datas = aluno['Datas']
    horas = aluno['Horas']
    preco = aluno['Preco']
    cumprimentos = cumprimento()
    p = Path(__file__).with_name('cobrancas.txt')
    with p.open('a', encoding='utf-8') as file:
        file.write(f'{cumprimentos}, {nome}. Segue o resumo das aulas de {
                   last_month_name}:\n\n')

        hours_count = 0
        for data, hora in zip(datas, horas):
            if data.month == last_month_number and data.day > 5 and data.year == datetime.now().year:
                pass
            elif data.month == datetime.now().month and data.day <= 5 and data.year == datetime.now().year:
                pass
            else:
                continue

            hours_count += hora

            file.write(f'{data.strftime("%d/%m")
                          } - {conditional_round(hora)}h\n')
        file.write('\n')

        valor_a_cobrar = hours_count * preco
        file.write(
            f'TOTAL: {conditional_round(hours_count)}h * R${preco:.2f} = R${valor_a_cobrar:.2f}\n')
        file.write(
            '\n###################################################################\n\n')
